import contextlib
import os
from io import BytesIO
from typing import Annotated

import win32com
import xlrd
from fastapi import Body, FastAPI, File, HTTPException, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from openpyxl import load_workbook
from src.api_llm import conversation_with_powerbi
from src.handle_excel_file import (
    XLS_SIGNATURE,
    MacroExecutionError,
    get_first_rows,
    get_xls_first_rows,
    inject_macro,
)
from src.prompts import _prompt_sys_template, format_data
from src.utils import get_substring

# Empty gen_py output directory
for file in os.listdir(win32com.__gen_path__):
    with contextlib.suppress(Exception):
        os.remove(os.path.join(win32com.__gen_path__, file))

app = FastAPI(
    title="API ChatBot",
    description="""**API ChatBot** est une IA qui permet de lire, traiter et modifier des fichiers Excel via l'injection de macros VBA générées par OpenAi.""",
    version="0.1.0",
    contact={
        "name": "YellowSys",
        "email": "mdougban@yellowsys.fr",
    },
)
origins = ["*"]

app.add_middleware(
    CORSMiddleware,
    allow_origins=origins,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)
app.mount("/data", StaticFiles(directory="data"), name="data")


@app.post("/handle_excel_file", tags=["QA"])
def handle_excel_file(
    username: Annotated[
        str,
        Body(
            title="A username",
            description="A valid username for the yellowsys llm api.",
        ),
    ],
    password: Annotated[
        str,
        Body(
            title="The user password",
            description="The user's password for using the yellowsys llm api.",
        ),
    ],
    query: Annotated[
        str,
        Body(
            title="The user query",
            description="The user query containing all the changes to be applied to the Excel file.",
        ),
    ],
    file: Annotated[
        UploadFile,
        File(
            title="The excel file to handle",
            description="The bytes object contains the Excel file you want to process.",
        ),
    ],
):
    """
    Generate VBA code using a yellowsys llm api and inject it into the excel file.

    Args:<br>
        `username` (str): A valid username for the yellowsys llm api.<br>
        `password` (str): The user's password for using the yellowsys llm api.<br>
        `query` (str): The user query containing all the changes to be applied to the Excel file.<br>
        `file` (bytes): The bytes object contains the Excel file you want to process.

    Returns:<br>
        A dict with the path to the new Excel file to download.<br>
    """
    # We assume that the data are on the first sheet.
    # And then select the frist 5 rows to pass to the prompt context
    # get file content
    file_content = file.file.read()
    # use xlrd to read xls file
    if file_content.startswith(XLS_SIGNATURE):
        wb = xlrd.open_workbook(file_contents=file_content)
        sheet_name = wb.sheet_names()[0]
        sheet = wb.sheet_by_name(sheet_name)
        first_rows = get_xls_first_rows(sheet)
    else:  # else try to use openpyxl for other format
        try:
            wb = load_workbook(filename=BytesIO(file_content))
        except Exception as e:
            raise HTTPException(
                status_code=500,
                detail="File is not supported! Only Excel xls, xlsx, xlsm, xltx and xltm files are supported",
            ) from e
        sheet_name = wb.sheetnames[0]
        sheet = wb[sheet_name]
        first_rows = get_first_rows(sheet)

    sys_role = _prompt_sys_template.format(
        sheet_name=sheet_name, first_rows=format_data(first_rows)
    )
    prompt = sys_role + """\n\n{history} \n\nHuman: {input}\n\nAssistant:"""
    response = conversation_with_powerbi(prompt, query, username, password)
    macro = get_substring(response["response"], start="Sub", end="End Sub")

    with open("./data/history.csv", "a") as f:
        f.write(f"{query}[SEP]{macro}[SEP]{response['response']}[EOR]\n")

    file_name = os.path.join("data", file.filename)
    with open(file_name, "wb") as f:
        f.write(file_content)
    try:
        xlsm_file = inject_macro(file_name, macro).replace("\\", "/")
        return {"url": xlsm_file}
    except MacroExecutionError as exc:
        raise HTTPException(
            status_code=500,
            detail=str(exc),
        ) from exc
